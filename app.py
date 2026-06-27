"""Konekxion backend.
Este módulo expone una aplicación Flask para gestionar contactos usando un
archivo Excel como almacenamiento persistente.
"""
# ---------------- Todas las librerias importadas ----------------
from flask import Flask, request, jsonify, render_template, redirect, url_for
from openpyxl import Workbook, load_workbook
import os
import re

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONTACTS_FILE = os.path.join(BASE_DIR, 'contacts.xlsx')
# Ruta absoluta al archivo de contactos. Evita errores cuando Flask se ejecuta desde un
# directorio distinto.
# Se usa para asegurar que el archivo Excel se abra correctamente incluso si la
# aplicación arranca desde una ruta de trabajo distinta.

# ---------------- Funcion para inicializar Excel y guardarlo  ----------------
def init_excel(file, headers):
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file)

# Inicializa el archivo Excel con encabezados si aún no existe.
init_excel(CONTACTS_FILE, [
    "Nombre completo",
    "Teléfono",
    "Correo",
    "Dirección",
    "Nota",
    "Categoría",
    "Favorito"
])

# ---------------- Página Principal ----------------
@app.route('/')
def index():
    return render_template('index.html')


# ---------------- LOGIN ----------------
@app.route('/login', methods=['POST'])
def login():
    """ Verifica que el usuario y contraseña sean correctos.
    Esto simula autenticación local. En producción debería usarse un sistema
    de usuarios y contraseñas seguro.
    """
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    if email == "admin" and password == "1234":
        return jsonify({"success": True}), 200

    return jsonify({
        "success": False,
        "error": "Usuario o contraseña inválidos"
    }), 400

# ---------------- Funcio para obtener valores de fila ----------------
def _row_value(row, index, default=''):
    """Get a row value safely, returning default when missing or None."""
    return row[index] if len(row) > index and row[index] is not None else default


# ---------------- Ruta MENÚ ----------------
@app.route('/menu')
def menu():
    """Muestra el menú principal con la lista de contactos cargada desde el archivo Excel.
       Si no hay muestra el archivo html """

    contactos = []

    if os.path.exists(CONTACTS_FILE):
        wb = load_workbook(CONTACTS_FILE)
        ws = wb.active

        # Leer Excel y enviar al HTML como lista de diccionarios
        for row in ws.iter_rows(min_row=2, values_only=True):
            contactos.append({
                "nombre": _row_value(row, 0),
                "telefono": _row_value(row, 1),
                "correo": _row_value(row, 2),
                "direccion": _row_value(row, 3),
                "nota": _row_value(row, 4),
                "categoria": _row_value(row, 5),
                "favorito": _row_value(row, 6, 'No')
            })

    return render_template(
        'menu.html',
        contactos=contactos
    )

# ---------------- Ruta FORMULARIO DE ACTUALIZACIÓN ----------------
@app.route('/update_contact_form')
def update_contact_form():
    """Muestra el formulario para actualizar un contacto existente, 
       cargando los datos actuales del contacto desde Excel."""
    
    # El nombre del contacto a actualizar se pasa como parámetro GET "old_nombre".
    old_nombre = request.args.get('old_nombre', '').strip()
    contacto = None

    if old_nombre and os.path.exists(CONTACTS_FILE):
        wb = load_workbook(CONTACTS_FILE)
        ws = wb.active

        # Buscar el contacto por nombre y cargar sus datos para mostrar en el formulario
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == old_nombre:
                contacto = {
                    'nombre': _row_value([cell.value for cell in row], 0),
                    'telefono': _row_value([cell.value for cell in row], 1),
                    'correo': _row_value([cell.value for cell in row], 2),
                    'direccion': _row_value([cell.value for cell in row], 3),
                    'nota': _row_value([cell.value for cell in row], 4),
                    'categoria': _row_value([cell.value for cell in row], 5),
                    'favorito': _row_value([cell.value for cell in row], 6, 'No')
                }
                break

    if not contacto:
        return redirect(url_for('menu'))

    return render_template('actualizar.html', contacto=contacto)

# ---------------- Ruta REPORTE ----------------
@app.route('/reporte')
def reporte():
    """Muestra una página con un reporte resumen del total de contactos y cuántos son favoritos.
       Los datos se calculan leyendo el archivo Excel."""
    total = 0
    favoritos = 0

    if os.path.exists(CONTACTS_FILE):
        wb = load_workbook(CONTACTS_FILE)
        ws = wb.active

        # Contar total de contactos y favoritos leyendo el Excel
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                total += 1
                if len(row) > 6 and row[6] == 'Sí':
                    favoritos += 1

    return render_template('reporte.html', total_contactos=total, favoritos=favoritos)

# ---------------- AGREGAR CONTACTO ----------------
@app.route('/add_contact')
def add_contact():
    """Muestra el formulario para agregar un nuevo contacto. No se necesitan datos previos, 
       solo se muestra el formulario vacío."""
    return render_template('add_contact.html')

# ---------------- Ruta GUARDAR CONTACTO ----------------
@app.route('/save_contact', methods=['POST'])
def save_contact():
    """Guarda un nuevo contacto enviado desde el formulario de agregar contacto en el archivo Excel.
       Valida los campos obligatorios, el formato del teléfono y correo antes de guardar."""
    
    nombre = request.form['nombre'].strip()
    telefono = request.form['telefono'].strip()
    correo = request.form['correo'].strip()
    direccion = request.form['direccion'].strip()
    nota = request.form.get('nota', '').strip()
    categoria = request.form.get('categoria', '').strip()
    favorito = "Sí" if 'favorito' in request.form else "No"

    # Validar campos obligatorios
    if not nombre:
        return "El nombre es obligatorio", 400
    if not telefono:
        return "El teléfono es obligatorio", 400
    if not correo:
        return "El correo es obligatorio", 400

    # Validar que el teléfono tenga exactamente 8 dígitos.
    if not re.fullmatch(r'\d{8}', telefono):
        return "El teléfono debe tener exactamente 8 dígitos", 400

    # Validar que el correo sea válido.
    if not re.fullmatch(r'^[\w\.-]+@[\w\.-]+\.\w+$', correo):
        return "Correo inválido", 400

    wb = load_workbook(CONTACTS_FILE)
    ws = wb.active

    ws.append([
        nombre,
        telefono,
        correo,
        direccion,
        nota,
        categoria,
        favorito
    ])

    wb.save(CONTACTS_FILE)
    return redirect(url_for('menu'))

# ---------------- Ruta ACTUALIZAR CONTACTO ----------------
@app.route('/update_contact', methods=['POST'])
def update_contact():
    """Actualiza un contacto existente en el archivo Excel con los datos enviados desde el formulario de actualización.
       Valida los campos obligatorios, el formato del teléfono y correo antes de guardar."""

    old_nombre = request.form['old_nombre']

    nombre = request.form['nombre']
    telefono = request.form['telefono']
    correo = request.form['correo']
    direccion = request.form['direccion']
    nota = request.form.get('nota', '')
    categoria = request.form.get('categoria', '').strip()
    favorito = "Sí" if 'favorito' in request.form else "No"

    # Validar campos obligatorios
    if not nombre.strip():
        return "El nombre es obligatorio", 400
    if not telefono.strip():
        return "El teléfono es obligatorio", 400
    if not correo.strip():
        return "El correo es obligatorio", 400

    # Validar que el teléfono tenga exactamente 8 dígitos.
    if not re.fullmatch(r'\d{8}', telefono):
        return "El teléfono debe tener exactamente 8 dígitos", 400

    # Validar que el correo sea válido.
    if not re.fullmatch(r'^[\w\.-]+@[\w\.-]+\.\w+$', correo):
        return "Correo inválido", 400

    wb = load_workbook(CONTACTS_FILE)
    ws = wb.active

    # Buscar el contacto por el nombre original (old_nombre) y actualizar sus datos en la fila correspondiente
    for row in ws.iter_rows(min_row=2):

        if row[0].value == old_nombre:

            row[0].value = nombre
            row[1].value = telefono
            row[2].value = correo
            row[3].value = direccion
            row[4].value = nota

            if len(row) > 5:
                row[5].value = categoria
            else:
                ws.cell(row=row[0].row, column=6, value=categoria)

            if len(row) > 6:
                row[6].value = favorito
            else:
                ws.cell(row=row[0].row, column=7, value=favorito)

            break

    wb.save(CONTACTS_FILE)
    return redirect(url_for('menu'))

# ---------------- Ruta ELIMINAR CONTACTOS ----------------
@app.route('/delete_contacts', methods=['POST'])
def delete_contacts():
    """Elimina uno o más contactos del archivo Excel según los números de fila recibidos desde el formulario 
       de selección.
       Valida que se reciban filas válidas y que el archivo exista antes de intentar eliminar los contactos."""
    data = request.get_json(silent=True) or {}
    rows = data.get('rows', [])

    # Validar que se recibieron filas para eliminar
    if not rows:
        return jsonify({
            "success": False,
            "error": "No se recibieron contactos para eliminar."
        }), 400
    # Validar que las filas sean números enteros válidos
    try:
        row_numbers = sorted({int(row) for row in rows}, reverse=True)
    except (TypeError, ValueError):
        return jsonify({
            "success": False,
            "error": "Datos de fila inválidos."
        }), 400
    # Validar que el archivo de contactos exista antes de intentar eliminar
    if not os.path.exists(CONTACTS_FILE):
        return jsonify({"success": False, "error": "El archivo de contactos no existe."}), 400

    wb = load_workbook(CONTACTS_FILE)
    ws = wb.active

    # Eliminar las filas correspondientes a los números recibidos, asegurando no eliminar la fila de encabezados
    for row_number in row_numbers:
        if row_number >= 2 and row_number <= ws.max_row:
            ws.delete_rows(row_number, 1)

    wb.save(CONTACTS_FILE)
    return jsonify({"success": True})

if __name__ == "__main__":
    app.run(debug=True)